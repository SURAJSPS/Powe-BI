#!/usr/bin/env python3
"""
Build RNK_Civil_Operations_AllPhases.xlsx — civil ops master workbook (Excel-only track).
Requires: openpyxl (install: pip install openpyxl, or use project .venv)
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).resolve().parent / "RNK_Civil_Operations_AllPhases.xlsx"

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
INSTRUCTIONS = """RNK Civil — Excel master workbook (all phases)

PHASE 1 — Master data: Setup_OTRules, Projects, Sites, Workers
PHASE 2 — Attendance: daily rows in Attendance (link WorkerID, ProjectCode, SiteCode)
PHASE 3 — Expenses: project costs, GST, approvals
PHASE 4 — Payroll: set period in Period_Control; review AttendanceSummary; create PayrollRuns + PayrollLines
PHASE 5 — Billing: Invoices + InvoiceLines (GST columns — confirm with your CA)
PHASE 6 — Dashboard: KPIs (uses Period_Control dates for MTD-style filters)

RULES
• Use Excel Tables only (do not delete header row). Add new rows inside the table.
• WorkerID / ProjectCode / SiteCode / InvoiceNo must stay unique and consistent.
• OneDrive/SharePoint: turn on version history; take monthly backup copies.
• This file is NOT a substitute for statutory payroll software — validate PF/ESI/PT/TDS with your CA.

OT & PAY (simplified)
• AttendanceSummary estimates pay using Period_Control dates. Monthly workers: EstBasePay uses gross/26 × paid working days cap (adjust policy in column note).
• Refine OT and statutory columns before real payroll runs.
"""


def style_header(ws, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table(ws, name: str, ref: str) -> None:
    tab = Table(displayName=name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def main() -> None:
    wb = Workbook()

    # --- Instructions ---
    ws0 = wb.active
    ws0.title = "Instructions"
    ws0["A1"] = INSTRUCTIONS
    ws0["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 100

    # --- Period_Control (Phase 4 / Dashboard) ---
    pc = wb.create_sheet("Period_Control", 1)
    pc["A1"], pc["B1"] = "PeriodLabel", "Apr-2026"
    pc["A2"], pc["B2"] = "PeriodStart", date(2026, 4, 1)
    pc["A3"], pc["B3"] = "PeriodEnd", date(2026, 4, 30)
    pc["A4"] = "Notes"
    pc["B4"] = "Change B2:B3 to filter AttendanceSummary & Dashboard period totals."
    pc["B2"].number_format = "yyyy-mm-dd"
    pc["B3"].number_format = "yyyy-mm-dd"
    style_header(pc, 2)
    pc.column_dimensions["A"].width = 16
    pc.column_dimensions["B"].width = 22

    # --- Setup_OTRules ---
    otr = wb.create_sheet("Setup_OTRules")
    hdr_otr = ["OTRuleID", "RuleName", "Multiplier", "MaxOTHrsPerDay", "Notes"]
    otr.append(hdr_otr)
    otr.append(["OT-STD", "Standard double rate", 2, 4, "Hours beyond 8 × (DailyRate/8)×Multiplier — refine in payroll"])
    style_header(otr, len(hdr_otr))
    add_table(otr, "tbl_OTRules", "A1:E2")

    # --- Projects ---
    prj = wb.create_sheet("Projects")
    hp = [
        "ProjectCode",
        "ProjectName",
        "ClientName",
        "StartDate",
        "EndDate",
        "Status",
        "Budget",
        "Description",
    ]
    prj.append(hp)
    prj.append(
        [
            "PRJ-001",
            "NH Road Package A",
            "NHAI",
            date(2026, 1, 1),
            date(2026, 12, 31),
            "Active",
            5000000,
            "Example active project",
        ]
    )
    prj.append(
        [
            "PRJ-002",
            "Urban Drainage",
            "Municipal Corp",
            date(2026, 3, 1),
            "",
            "Planning",
            1200000,
            "Second project",
        ]
    )
    style_header(prj, len(hp))
    add_table(prj, "tbl_Projects", f"A1:{get_column_letter(len(hp))}3")
    dv_status = DataValidation(
        type="list",
        formula1='"Planning,Active,On Hold,Completed,Archived"',
        allow_blank=True,
    )
    dv_status.add(f"F2:F500")
    prj.add_data_validation(dv_status)

    # --- Sites ---
    st = wb.create_sheet("Sites")
    hs = ["SiteCode", "ProjectCode", "SiteName", "Location", "Active"]
    st.append(hs)
    st.append(["S-001", "PRJ-001", "Km 12–15 Site Office", "NH-44 near km 12", "Yes"])
    st.append(["S-002", "PRJ-001", "Batch Plant", "Off-site batching", "Yes"])
    style_header(st, len(hs))
    add_table(st, "tbl_Sites", f"A1:{get_column_letter(len(hs))}3")
    dv_yes = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_yes.add("E2:E500")
    st.add_data_validation(dv_yes)

    # --- Workers ---
    wk = wb.create_sheet("Workers")
    hw = [
        "WorkerID",
        "FullName",
        "Phone",
        "PayType",
        "DailyRate",
        "MonthlyGross",
        "OTRuleID",
        "RateEffectiveDate",
        "Skill",
        "IndiaState",
        "PAN",
        "Active",
        "PrimaryProjectCode",
    ]
    wk.append(hw)
    wk.append(
        [
            "W-001",
            "Ramesh Kumar",
            "9876500001",
            "Daily",
            800,
            "",
            "OT-STD",
            date(2026, 4, 1),
            "Mason",
            "Maharashtra",
            "",
            "Yes",
            "PRJ-001",
        ]
    )
    wk.append(
        [
            "W-002",
            "Suresh Patil",
            "9876500002",
            "Monthly",
            "",
            22000,
            "OT-STD",
            date(2026, 4, 1),
            "Supervisor",
            "Maharashtra",
            "ABCDE1234F",
            "Yes",
            "PRJ-001",
        ]
    )
    wk.append(
        [
            "W-003",
            "Vendor Labour A",
            "",
            "Daily",
            700,
            "",
            "OT-STD",
            date(2026, 4, 1),
            "Helper",
            "Maharashtra",
            "",
            "Yes",
            "PRJ-002",
        ]
    )
    style_header(wk, len(hw))
    add_table(wk, "tbl_Workers", f"A1:{get_column_letter(len(hw))}4")
    dv_pt = DataValidation(type="list", formula1='"Daily,Monthly"', allow_blank=False)
    dv_pt.add("D2:D2000")
    wk.add_data_validation(dv_pt)
    dv_act = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv_act.add("L2:L2000")
    wk.add_data_validation(dv_act)
    for col in ["E", "F"]:
        for r in range(2, 5):
            wk[f"{col}{r}"].number_format = "#,##0.00"

    # --- Attendance ---
    att = wb.create_sheet("Attendance")
    ha = [
        "Date",
        "WorkerID",
        "ProjectCode",
        "SiteCode",
        "Status",
        "NormalHrs",
        "OTHrs",
        "Notes",
    ]
    att.append(ha)
    rows_att = [
        [date(2026, 4, 1), "W-001", "PRJ-001", "S-001", "Present", 8, 2, ""],
        [date(2026, 4, 1), "W-002", "PRJ-001", "S-001", "Present", 8, 0, ""],
        [date(2026, 4, 2), "W-001", "PRJ-001", "S-001", "Present", 8, 0, "Full day"],
        [date(2026, 4, 2), "W-002", "PRJ-001", "S-001", "Absent", 0, 0, "Leave"],
    ]
    for row in rows_att:
        att.append(row)
    style_header(att, len(ha))
    add_table(att, "tbl_Attendance", f"A1:{get_column_letter(len(ha))}{1 + len(rows_att)}")
    dv_st = DataValidation(
        type="list",
        formula1='"Present,Absent,Leave,Holiday,Half-day"',
        allow_blank=True,
    )
    dv_st.add("E2:E10000")
    att.add_data_validation(dv_st)
    for r in range(2, 2 + len(rows_att)):
        att.cell(row=r, column=1).number_format = "yyyy-mm-dd"

    # --- Expenses ---
    ex = wb.create_sheet("Expenses")
    he = [
        "ExpenseDate",
        "ProjectCode",
        "Category",
        "Amount",
        "GSTAmount",
        "Vendor",
        "Approved",
        "BillLinkOrRef",
        "Notes",
    ]
    ex.append(he)
    ex.append(
        [
            date(2026, 4, 3),
            "PRJ-001",
            "Petty cash",
            3500,
            0,
            "Local vendor",
            "Yes",
            "INV/PC/12",
            "Site consumables",
        ]
    )
    ex.append(
        [
            date(2026, 4, 4),
            "PRJ-001",
            "Fuel",
            8200,
            1476,
            "HP Petrol",
            "No",
            "",
            "Awaiting approval",
        ]
    )
    style_header(ex, len(he))
    add_table(ex, "tbl_Expenses", f"A1:{get_column_letter(len(he))}3")
    dv_ap = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv_ap.add("G2:G5000")
    ex.add_data_validation(dv_ap)
    for r in range(2, 4):
        ex.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ex.cell(row=r, column=4).number_format = "#,##0.00"
        ex.cell(row=r, column=5).number_format = "#,##0.00"

    # --- AttendanceSummary (formulas; Period_Control drives dates) ---
    sm = wb.create_sheet("AttendanceSummary")
    hsm = [
        "WorkerID",
        "FullName",
        "PayType",
        "DailyRate",
        "MonthlyGross",
        "OTRuleID",
        "PaidDays",
        "TotalNormalHrs",
        "TotalOTHrs",
        "Multiplier",
        "EstOTPay",
        "EstBasePay",
        "EstGrossBeforeStatutory",
        "Notes",
    ]
    sm.append(hsm)
    style_header(sm, len(hsm))
    # One formula row per sample worker — user can extend table downward
    workers_for_summary = ["W-001", "W-002", "W-003"]
    for i, wid in enumerate(workers_for_summary, start=2):
        sm.cell(row=i, column=1, value=wid)
        sm.cell(
            row=i,
            column=2,
            value=f'=IFERROR(XLOOKUP([@[WorkerID]],tbl_Workers[WorkerID],tbl_Workers[FullName]),"")',
        )
        sm.cell(
            row=i,
            column=3,
            value=f'=IFERROR(XLOOKUP([@[WorkerID]],tbl_Workers[WorkerID],tbl_Workers[PayType]),"")',
        )
        sm.cell(
            row=i,
            column=4,
            value=f'=IFERROR(XLOOKUP([@[WorkerID]],tbl_Workers[WorkerID],tbl_Workers[DailyRate]),0)',
        )
        sm.cell(
            row=i,
            column=5,
            value=f'=IFERROR(XLOOKUP([@[WorkerID]],tbl_Workers[WorkerID],tbl_Workers[MonthlyGross]),0)',
        )
        sm.cell(
            row=i,
            column=6,
            value=f'=IFERROR(XLOOKUP([@[WorkerID]],tbl_Workers[WorkerID],tbl_Workers[OTRuleID]),"")',
        )
        sm.cell(
            row=i,
            column=7,
            value=(
                '=COUNTIFS(tbl_Attendance[WorkerID],[@[WorkerID]],tbl_Attendance[Date],">="&Period_Control!$B$2,'
                'tbl_Attendance[Date],"<="&Period_Control!$B$3,tbl_Attendance[Status],"Present")'
            ),
        )
        sm.cell(
            row=i,
            column=8,
            value=(
                '=SUMIFS(tbl_Attendance[NormalHrs],tbl_Attendance[WorkerID],[@[WorkerID]],'
                'tbl_Attendance[Date],">="&Period_Control!$B$2,tbl_Attendance[Date],"<="&Period_Control!$B$3)'
            ),
        )
        sm.cell(
            row=i,
            column=9,
            value=(
                '=SUMIFS(tbl_Attendance[OTHrs],tbl_Attendance[WorkerID],[@[WorkerID]],'
                'tbl_Attendance[Date],">="&Period_Control!$B$2,tbl_Attendance[Date],"<="&Period_Control!$B$3)'
            ),
        )
        sm.cell(
            row=i,
            column=10,
            value=f'=IFERROR(XLOOKUP([@[OTRuleID]],tbl_OTRules[OTRuleID],tbl_OTRules[Multiplier]),0)',
        )
        # OT pay: (OTHrs) * (DailyRate/8) * Multiplier  [simplified; assumes hourly OT from daily rate]
        sm.cell(
            row=i,
            column=11,
            value="=[@[TotalOTHrs]]*([@[DailyRate]]/8)*[@[Multiplier]]",
        )
        sm.cell(
            row=i,
            column=12,
            value=(
                '=IF([@[PayType]]="Daily",[@[PaidDays]]*[@[DailyRate]],'
                'IF([@[PayType]]="Monthly",[@[MonthlyGross]]/26*MIN([@[PaidDays]],26),0))'
            ),
        )
        sm.cell(
            row=i,
            column=13,
            value="=[@[EstBasePay]]+[@[EstOTPay]]",
        )
        sm.cell(row=i, column=14, value="Review LOP/statutory before paying")

    add_table(sm, "tbl_AttendanceSummary", f"A1:{get_column_letter(len(hsm))}{1 + len(workers_for_summary)}")
    for c in range(4, 14):
        for r in range(2, 2 + len(workers_for_summary)):
            sm.cell(row=r, column=c).number_format = "#,##0.00"
    sm.column_dimensions["B"].width = 22

    # --- PayrollRuns ---
    pr = wb.create_sheet("PayrollRuns")
    hpr = ["RunID", "PeriodLabel", "PeriodStart", "PeriodEnd", "Status", "ApprovedBy", "ApprovedDate", "Notes"]
    pr.append(hpr)
    pr.append(
        [
            "RUN-202604",
            "Apr-2026",
            date(2026, 4, 1),
            date(2026, 4, 30),
            "Draft",
            "",
            "",
            "Lock after CA review",
        ]
    )
    style_header(pr, len(hpr))
    add_table(pr, "tbl_PayrollRuns", f"A1:{get_column_letter(len(hpr))}2")
    dv_prs = DataValidation(type="list", formula1='"Draft,Approved,Paid,Locked"', allow_blank=True)
    dv_prs.add("E2:E500")
    pr.add_data_validation(dv_prs)
    pr["C2"].number_format = pr["D2"].number_format = "yyyy-mm-dd"

    # --- PayrollLines (India: earnings + deductions placeholders) ---
    pl = wb.create_sheet("PayrollLines")
    hpl = [
        "RunID",
        "WorkerID",
        "Component",
        "Amount",
        "StatutoryTag",
        "Notes",
    ]
    pl.append(hpl)
    sample_lines = [
        ["RUN-202604", "W-001", "Gross_Est", 0, "Earning", "=EstGross from AttendanceSummary"],
        ["RUN-202604", "W-001", "PF_Employee", 0, "Deduction", "Enter per CA"],
        ["RUN-202604", "W-002", "Gross_Est", 0, "Earning", ""],
    ]
    for row in sample_lines:
        pl.append(row)
    style_header(pl, len(hpl))
    add_table(pl, "tbl_PayrollLines", f"A1:{get_column_letter(len(hpl))}{1 + len(sample_lines)}")
    for r in range(2, 2 + len(sample_lines)):
        pl.cell(row=r, column=4).number_format = "#,##0.00"

    # --- Invoices ---
    inv = wb.create_sheet("Invoices")
    hin = [
        "InvoiceNo",
        "InvoiceDate",
        "ProjectCode",
        "ClientName",
        "ClientGSTIN",
        "SubTotal",
        "CGST",
        "SGST",
        "IGST",
        "Total",
        "RetentionPct",
        "Status",
        "Notes",
    ]
    inv.append(hin)
    inv.append(
        [
            "INV-2026-001",
            date(2026, 4, 5),
            "PRJ-001",
            "NHAI",
            "27AAAAA0000A1Z5",
            100000,
            9000,
            9000,
            0,
            118000,
            10,
            "Sent",
            "Example GST split — verify with CA",
        ]
    )
    style_header(inv, len(hin))
    add_table(inv, "tbl_Invoices", f"A1:{get_column_letter(len(hin))}2")
    inv["B2"].number_format = "yyyy-mm-dd"
    for col in range(6, 11):
        inv.cell(row=2, column=col).number_format = "#,##0.00"

    # --- InvoiceLines ---
    il = wb.create_sheet("InvoiceLines")
    hil = ["InvoiceNo", "LineNo", "Description", "HSN_SAC", "TaxableAmount", "GST_Rate_Pct", "Notes"]
    il.append(hil)
    il.append(["INV-2026-001", 1, "Mobilization & site establishment", "9954", 100000, 18, ""])
    style_header(il, len(hil))
    add_table(il, "tbl_InvoiceLines", f"A1:{get_column_letter(len(hil))}2")
    il["E2"].number_format = "#,##0.00"

    # --- Dashboard ---
    db = wb.create_sheet("Dashboard")
    db["A1"] = "RNK Civil — KPIs (Phase 6)"
    db["A1"].font = Font(bold=True, size=14)
    db["A3"] = "Active projects"
    db["B3"] = '=COUNTIFS(tbl_Projects[Status],"Active")'
    db["A4"] = "Total projects"
    db["B4"] = "=COUNTA(tbl_Projects[ProjectCode])"
    db["A5"] = "Active workers (Yes)"
    db["B5"] = '=COUNTIFS(tbl_Workers[Active],"Yes")'
    db["A6"] = "Period expenses (Approved=Yes)"
    db["B6"] = (
        '=SUMIFS(tbl_Expenses[Amount],tbl_Expenses[ExpenseDate],">="&Period_Control!$B$2,'
        'tbl_Expenses[ExpenseDate],"<="&Period_Control!$B$3,tbl_Expenses[Approved],"Yes")'
    )
    db["A7"] = "Period expense rows (all)"
    db["B7"] = (
        '=SUMIFS(tbl_Expenses[Amount],tbl_Expenses[ExpenseDate],">="&Period_Control!$B$2,'
        'tbl_Expenses[ExpenseDate],"<="&Period_Control!$B$3)'
    )
    db["A8"] = "Attendance Present days (all workers, period)"
    db["B8"] = (
        '=COUNTIFS(tbl_Attendance[Date],">="&Period_Control!$B$2,tbl_Attendance[Date],"<="&Period_Control!$B$3,'
        'tbl_Attendance[Status],"Present")'
    )
    db["A9"] = "Sum Est Gross (AttendanceSummary)"
    db["B9"] = "=SUM(tbl_AttendanceSummary[EstGrossBeforeStatutory])"
    db.column_dimensions["A"].width = 48
    db.column_dimensions["B"].width = 18
    for r in range(3, 10):
        db.cell(row=r, column=2).number_format = "#,##0.00"

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
